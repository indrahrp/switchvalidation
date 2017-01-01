import openpyxl,pprint
from load_switchport6 import *
from collections import OrderedDict
import getopt, sys


class server(object):
        def __init__(self,name):
                self.name = name


        def __str__(self):
                return self.name
        def __repr__(self):
                return "'"+self.name+"'"
	
		#def __eq__(self, other): 
		#	#return self.__dict__ == other.__dict__
		#	return self.name == other.name
			
def Load_Master_Sheet (server_detail): 
	print('Opening Workbook ...')
	wb = openpyxl.load_workbook('tdn-hzl-2016-05-15 taken sept2.xlsx')
	#wb = openpyxl.load_workbook('mastersheet.xlsx')
	sheet = wb.get_sheet_by_name('Z2 detail')
	for row in range(2, sheet.max_row):
		server_name = sheet['A' + str(row)].value   
		systemPort	= sheet['E'	+ str(row)].value
		switch_name	= sheet['F'	+ str(row)].value
		switch_port	= sheet['H' + str(row)].value
		logical_net	= sheet['I' + str(row)].value
		
		svr=server(server_name)
		server_detail.setdefault(svr,{})
		server_detail [svr]['systemPort']=systemPort
		server_detail [svr]['switch_name']=switch_name
		server_detail [svr]['switch_port']=switch_port
		server_detail [svr]['logical_net']=logical_net
		print (' server names ' + server_name + ' system port ' + systemPort + ' switch name ' + server_detail [svr]['switch_name'] + ' switch port '+server_detail [svr]['switch_port']   )
		print "SERVER DETAIL " + str(server_detail)

def printinfoone (server_detail,server_name):
	for server in server_detail:
		print ('Server Name all : ' + str(server) + ' ' + server_detail[server]['switch_name'])
		

def printinfoall (server_detail):
	#pprint.pformat(server_detail)
	for theserver,serverinfo in (server_detail.items()):
		#if serverinfo['systemPort'] is None:
		#	print "None " 
			#print ('Server Name : ' + theserver + ' ' + theserver[server]['systemPort']  + ' ' + 
			#server[server]['switch_name'] + ' ' + server.get([server]['switch_port'])) 
			print (str(theserver) + ' '+ serverinfo['systemPort'] + ' ' + serverinfo['switch_name'] + ' '+ serverinfo['switch_port']+ ' '+serverinfo['logical_net'])
	        #print (str(theserver) + ' ' + serverinfo.get('systemPort'),None)
		    #print (str(theserver) + '2' )

def MatchSwitchPort(tableinfo,server_detail,queryonesvr):
	
	for svrname,svrinfo in server_detail.items():	
		#if reportsvr <> 'all' and str(svrname).lower <> reportsvr.lower: 
		if str(svrname) != str(queryonesvr) and str(queryonesvr) != 'all':
			dummy=0
			#print (" break the loop for server : " + str(reportsvr) + " and svr name is " + str(svrname))
			#print (" break the loop for server : " + str(queryonesvr))
			
		else:
			foutput = open('Z2a_Results','a')
			print('Looking for server '+ str(svrname) + ' ' + svrinfo['systemPort'] + ' '+ svrinfo['switch_name'] + ' '+svrinfo['switch_port'])
			#print('server '+ str(svrname) + ' '+ svrinfo['switch_name'] + ' '+svrinfo['switch_port']) 
			for switchname,switchinfo in tableinfo.items():
				# for debug print('(switchname '+switchname + ' switchport ' + switchinfo[1])
				if svrinfo['switch_name'] == switchinfo[0]:
					switchprt='E'+ str(switchinfo[1])
					#for debug print (svrinfo['switch_name'] + ' ' + svrinfo['switch_port'] + ' ' + switchinfo[0] + ' switchprt ' + switchprt )
					if re.sub('/0','/',svrinfo['switch_port']) == switchprt:
						#for debug print ('match ' + svrinfo['switch_name'] + ' '+ switchname + ' '+ svrinfo['switch_port'] + ' ' + switchinfo[1])
						Native,AllowedVlan=GetVlanDefinition(svrinfo['logical_net'])
						print (Native+ ' '+ AllowedVlan +'  switch native ' + switchinfo[5] + ' swith allow ' + switchinfo[7])
						if switchinfo[5].endswith(Native): 
						    Mesg="Correct Native Vlan"
							#print('Correct Native Vlan')
						else:
						    Mesg='Incorrect Native Vlan'
							#print('Incorrect Native Vlan')
						if svrinfo['logical_net'] == 'TKR' or svrinfo['logical_net'] == 'EXCH' or svrinfo['logical_net'] == 'ETG' or svrinfo['logical_net'] == 'FUNDIST':
							if re.sub('switchport.*vlan ','',switchinfo[7]) in AllowedVlan:
								Mesg1='Correct Allowed Vlan'
								print('Correct Allowed Vlan')
							else:
								Mesg1='Incorrect Allowed Vlan'
								print('Incorrect Allowed Vlan ')
							
						elif switchinfo[7].endswith(AllowedVlan):
						#if switchinfo[7].find(AllowedVlan):
							print('Correct Allowed Vlan')
							Mesg1='Correct Allowed Vlan'
						else:
							Mesg1='Incorrect Allowed Vlan'
							print('Incorrect Allowed Vlan ')
							
						#print ( 'server description ' + str(svrname)+' switch description ' + switchinfo[2])
						if str(svrname).lower() in switchinfo[2].lower():
						#	print('correct description')
						    Mesg2='correct description'
						else:
							print('incorrect description')
							Mesg2='incorrect description'
						
						#writetoexcel(str(svrname),svrinfo['switch_name'],svrinfo['switch_port'],switchinfo['switch_name'),switchinfo[2],switchprt,switchinfo[5],switchinfo[7])
						foutput.write(str(svrname)+ ';'+ svrinfo['systemPort']+';'+ svrinfo['switch_name'] + ';' + svrinfo['switch_port'] + ';' + svrinfo['logical_net'] +
						';' +switchinfo[0] + ';' + switchinfo[2] + ';' + switchprt + ';'+ switchinfo[5] + ';' +	 switchinfo[7] +';'+ Mesg +';' + Mesg1 + ';' + Mesg2+';' + Native + ';' +AllowedVlan + '\n')
					
						print(str(svrname)+ ' '+ svrinfo['systemPort']+' '+ svrinfo['switch_name'] + ' ' + svrinfo['switch_port'] + ' ' + svrinfo['logical_net'] +
						' ' +switchinfo[0] + ' ' + switchinfo[2] + ' ' + switchprt + ' '+ switchinfo[5] + ' ' +	 switchinfo[7] +' '+ Mesg +' ' + Mesg1 + ' ' + Mesg2+' ' + Native + ' ' +AllowedVlan + '\n')
					
					    
#writetoexcel(svrname,svr_switch_name,,svr_switch_port,switch_switch_name,switch_description,switch_port,switch_native,switch_allowed)


			
						
def GetVlanDefinition(logical_net):
	LogicalNetNativeVlan= ({'ARBI40':'901', 'ARBI10':'901','PROD':'2201','BUILD':'2230','CHASSIS':'2200','ETG':'901','EXCH':'901','EXT':'2230','EXTHYPER':'2200',
	'FEED':'901','FEPC':'901','FUNDIST':'901','GTDL':'901','HOST':'2241','HOSTHYPER':'2200','HOSTREP':'2243','IDMZ':'2220','LOG':'901','PRODHYPER':'2200','PRODPDNHYPER':'2200',
	'TKR':'901','TPFUN':'901','WEB':'2210','WEBHOST':'2242','WEBHYPER':'2200','PRODPDNLBTEMP':'4000','VDITERM':'2231'})
	
	LogicalNetAllowedVlan=({'ARBI40':'1701','ARBI10':'1703','PROD':'2201-2209',
	'BUILD':'2200-2202,2209-2210,2230','CHASSIS':'2200-2219,2230-2239,2244-2246','ETG':'815,816','EXCH':'1501-1502,1601','EXT':'2230-2239','EXTHYPER':'2200,2230-2239',
	'FEED':'1800','FEPC':'1801','FUNDIST':'258,13','GTDL':'1802','HOST':'2243','HOSTHYPER':'2200,2241,2243','HOSTREP':'2243','IDMZ':'2200,2220,2280,2295','IDMZHYPER':'2200,2220,2280,2295','LOG':'1804','PRODHYPER':'2200-2209',
	'PRODPDN':'2201-2209,2244-2246','PRODPDNHYPER':'2200-2209,2244-2246','PRODWEBHYPER':'2200-2209,2210-2219','TKR':'1511-1519,1611-1619','TPFUN':'1805','WEB':'2210-2219','WEBHOST':'2242,2562','WEBHYPER':'2200,2210-219',
	'WEB':'2210-2219','WEBHOST':'2242','WEBHYPER':'2200,2210-2219','PRODPDNLBTEMP':'4000','VDITERM':'2231'})
	
	return(LogicalNetNativeVlan[logical_net],LogicalNetAllowedVlan[logical_net])
	
			
			
			
server_detail=OrderedDict()
Load_Master_Sheet(server_detail)
#print str(server_detail)
#print server_detail
#printinfoone(server_detail,server('all'))
#printinfoall (server_detail)

SwitchConfig=['C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG2-SWA1.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG2-SWA2.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG2-SWA3.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG2-SWA4.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG7-SWA1.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG7-SWA2.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG7-SWA3.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG7-SWA5.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG7-SWA6.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG8-SWA1.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNAGG8-SWA2.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH2-SWA1.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH2-SWA2.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH5-SWA1.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH5-SWA2.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH7-SWA1.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH7-SWA2.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH8-SWA1.txt',
'C:\Python27\mine\switch\HZL-COLO-PRODTDNDH8-SWA2.txt'
]

def usage():
	print "\n -h for help "
	print " -q query detail of a server"
	print " -R for query detail of all server"


tableinfo = {}
for Switchfile in SwitchConfig:
	SwitchName=switchNameRegex.search(Switchfile).group(0)
	print('Switch Name :' +  SwitchName )
	Res=ReadSwitchConfigFromFile(Switchfile)
	ParseSwitch(Res,Switchfile,SwitchName,tableinfo)

	
#queryonesvr=server('marketz2')	
#MatchSwitchPort(tableinfo,server_detail,queryonesvr)


def main():
    try:
        opts, args = getopt.getopt(sys.argv[1:], "hq:R")
		#opts, args = getopt.getopt(sys.argv[1:], "ho:v", ["help", "output="])
    except getopt.GetoptError as err:
        # print help information and exit:
        print str(err)  # will print something like "option -a not recognized"
        usage()
        sys.exit(2)
		
    for o, a in opts:
		if o == "-h":
			usage()
			sys.exit(0)
		elif o == "-q":
			queryonesvr=server(a)
			MatchSwitchPort(tableinfo,server_detail,queryonesvr)
			#sys.exit(0)
		elif o == "-R":
			queryonesvr=server('all')
			MatchSwitchPort(tableinfo,server_detail,queryonesvr)
		else:
			assert False, "unhandled option"
		 
		#lif o == "-a
		#adding account=true
		#	adding account=true
        #else:
        #    assert False, "unhandled option"


if __name__ == "__main__":
	main()

